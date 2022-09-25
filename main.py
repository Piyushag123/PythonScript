from openpyxl import *
import os.path
import requests
import json





response_API = requests.get('https://jsonplaceholder.typicode.com/users')
data = response_API.text
parse_json = json.loads(data)




def fun():
    wb = Workbook()

# open workbook
    ws = wb.active

# modify the desired cell
    ws.title = "Ram"
    a = ["EmpId", "Name", "UserName", "Website", "Verification_Status"]
    for i in range(5):
        ws.cell(row=1, column=i+1, value=a[i])
#ws["A1"] = "Full Name"
    print("Hii Welcome to the Excel Dashboard")

#total = int(input("Enter the number of details u want to add"))
    #i = 1
    #while True:
        #empid = int(input("Enter the emp id of user"))

        #firstname = input("Enter the first name of user")
        #lastname = input("Enter the last name of user")
        #address = input("Enter the emp id of user")
        #ws.cell(row=i+1, column=1, value=empid)
        #ws.cell(row=i+1, column=2, value=firstname)
        #ws.cell(row=i + 1, column=3, value=lastname)
        #ws.cell(row=i + 1, column=4, value=address)
        #i = i+1
        #ch = input("Do You Want To add more details ?yes/no")
        #if ch == "no":
            #break

# save the file
    for i in parse_json:
        a = i['id'], i['name'], i['username'], i['website']
        ws.append(a)

    wb.save(filename="output.xlsx")


def append():
    wc = load_workbook("output.xlsx")
    ws = wc.active
    curr_cell = ws.max_row
    #print(curr_cell)
    print("You are inserting into existing excel file")
    i = curr_cell
    while True:
        empid = int(input("Enter the emp id of user"))

        firstname = input("Enter the first name of user")
        lastname = input("Enter the last name of user")
        address = input("Enter the emp id of user")
        ws.cell(row=i + 1, column=1, value=empid)
        ws.cell(row=i + 1, column=2, value=firstname)
        ws.cell(row=i + 1, column=3, value=lastname)
        ws.cell(row=i + 1, column=4, value=address)
        i = i + 1
        ch = input("Do You Want To add more details ?yes/no")
        if ch == "no":
            break

    wc.save(filename="output.xlsx")


def validation():
    wc = load_workbook("output.xlsx")
    ws = wc.active
    wd = load_workbook("validData.xlsx")
    wl = wd.active
    curr_cell = ws.max_row
    curr_cell1 = wl.max_row
    a = []
    for i in range(2, curr_cell1+1):
        a.append(wl.cell(row=i, column=1).value
                 )


    #for row in wc.iter_rows(min_row=1, min_col=1, max_row=wc.max_row, max_col=5):
        #for cell in row:
            #for i in a:
                #if(cell.value == i):
                    #print("hii")

    for i in a:
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=curr_cell, max_col=5):
            if(row[0].value == i):
                row[4].value = "Approved"
                break
        pass

    wc.save(filename="output.xlsx")










if os.path.isfile('output.xlsx'):
    append()

else:
    fun()


validation()